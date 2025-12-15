import pandas as pd
import json

def excel_para_json_alternativo(arquivo_excel, arquivo_json):
    
    try:
        from openpyxl import load_workbook
        
        # Carrega o workbook
        wb = load_workbook(arquivo_excel, data_only=True)
        ws = wb.active
        
        dados = []
        estado_atual = None
        
        for row in ws.iter_rows():
            # Extrai valores da linha
            valores = []
            for cell in row:
                if cell.value is not None:
                    valores.append(str(cell.value).strip())
                else:
                    valores.append("")
            
            # Remove colunas vazias no fim
            while valores and valores[-1] == "":
                valores.pop()
            
            if not valores:
                continue
            
            # Se a linha tem apenas 1 valor → é o ESTADO
            if len(valores) == 1:
                estado_atual = valores[0]
                print(f"📍 Processando estado: {estado_atual}")
                continue
            
            # Se a linha tem pelo menos 4 valores → é registro válido
            if estado_atual and len(valores) >= 4:
                registro = {
                    "estado": estado_atual,
                    "UF": valores[0].replace("\n", " "),
                    "codigo": valores[1].replace("\n", " ").zfill(2),
                    "OM_titulo": valores[2].replace("\n", " "),
                    "OM_descricao": valores[3].replace("\n", " ")
                }
                dados.append(registro)
        
        # Salva em JSON
        with open(arquivo_json, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)
        
        print(f"✅ JSON gerado com sucesso!")
        print(f"📄 Arquivo: {arquivo_json}")
        print(f"📊 Total de registros: {len(dados)}")
        
        return True
        
    except ImportError:
        print("❌ openpyxl não está instalado. Instale com: pip install openpyxl")
        return False
    except Exception as e:
        print(f"❌ Erro ao processar arquivo: {e}")
        return False


if __name__ == "__main__":
    arquivo_excel = "Localidades.xlsx"
    arquivo_json = "tabela.json"
    
    print("🚀 Iniciando conversão Excel → JSON")
    print("-" * 50)
    
    excel_para_json_alternativo(arquivo_excel, arquivo_json)